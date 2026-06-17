"""Seed the database from the existing CD Engineering proposal workbook.

Usage:
    python import_excel.py "D:/My Drive/5. DESIGN DOCS/Project Proposal/2025 - Proposal Submital Fee .xlsm"
"""
from __future__ import annotations

import sys
import warnings
from datetime import datetime, date

warnings.filterwarnings("ignore")

import openpyxl

from app import app
from extensions import db
from models import Client, Architect, Company, Proposal, User


EXCEL_ERROR_TOKENS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in EXCEL_ERROR_TOKENS:
        return None
    return s


def _date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _float(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace(",", "").replace("₱", "").strip()
        return float(s)
    except ValueError:
        return None


_company_cache: dict[str, Company] = {}
_architect_cache: dict[str, Architect] = {}
_client_cache: dict[str, Client] = {}


def _norm(s):
    return s.casefold().strip() if s else ""


def _get_or_create_company(name, address=None):
    name = _str(name)
    if not name:
        return None
    key = _norm(name)
    if key in _company_cache:
        c = _company_cache[key]
        if address and not c.address:
            c.address = _str(address)
        return c
    company = Company(name=name, address=_str(address))
    db.session.add(company)
    db.session.flush()
    _company_cache[key] = company
    return company


def _get_or_create_architect(name, company=None):
    name = _str(name)
    if not name:
        return None
    key = _norm(name)
    if key in _architect_cache:
        a = _architect_cache[key]
        if company and not a.company:
            a.company = company
        return a
    architect = Architect(name=name, company=company)
    db.session.add(architect)
    db.session.flush()
    _architect_cache[key] = architect
    return architect


def _get_or_create_client(name):
    name = _str(name)
    if not name:
        return None
    key = _norm(name)
    if key in _client_cache:
        return _client_cache[key]
    client = Client(name=name)
    db.session.add(client)
    db.session.flush()
    _client_cache[key] = client
    return client


def _warm_caches():
    for c in Company.query.all():
        _company_cache[_norm(c.name)] = c
    for a in Architect.query.all():
        _architect_cache[_norm(a.name)] = a
    for c in Client.query.all():
        _client_cache[_norm(c.name)] = c


def import_workbook(path):
    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path, keep_vba=False, data_only=True)
    if "Summary" not in wb.sheetnames:
        print("No 'Summary' sheet found — aborting.")
        return

    ws = wb["Summary"]
    # Expected header row 2:
    # Link | Date Received | Category | Project Title | Owner | Location |
    # Price | Architect in Charge | Company | Remark | Price Approved |
    # Date Approved | Payed | Payable | Payment | Project # | Update
    created = 0
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not any(row):
            continue

        (
            link, date_received, category, project_title, owner, location,
            price, architect_name, company_name, remark, price_approved,
            date_approved, payed, payable, payment, project_number, update,
            *_rest
        ) = (list(row) + [None] * 17)[:17]

        title = _str(project_title)
        if not title:
            skipped += 1
            continue

        # Build relations
        company = _get_or_create_company(company_name)
        architect = _get_or_create_architect(architect_name, company)
        client = _get_or_create_client(owner)

        # Skip duplicates by project number
        pnum = _str(project_number)
        if pnum and Proposal.query.filter_by(project_number=f"P{pnum}").first():
            skipped += 1
            continue

        price_val = _float(price)
        approved_val = _float(price_approved)
        payed_val = _float(payed) or 0

        # Decide status from sheet data
        remark_str = (_str(remark) or "").lower()
        if "approved" in remark_str:
            status = "Paid" if payed_val and approved_val and payed_val >= approved_val else "Approved"
        elif approved_val:
            status = "Approved"
        else:
            status = "Submitted"

        proposal = Proposal(
            project_number=f"P{pnum}" if pnum else None,
            date_received=_date(date_received) or date.today(),
            date_proposed=_date(date_received) or date.today(),
            project_title=title,
            category=_str(category) or "Structural Design",
            location=_str(location),
            client=client,
            architect=architect,
            price=price_val if isinstance(price_val, float) else None,
            price_approved=approved_val,
            payed=payed_val,
            payment_method=_str(payment),
            date_approved=_date(date_approved),
            status=status,
            remarks=_str(remark),
        )
        db.session.add(proposal)
        created += 1

        if created % 50 == 0:
            db.session.commit()
            print(f"  …committed {created} proposals so far")

    db.session.commit()
    print(f"\nDone. Imported {created} proposals, skipped {skipped} rows.")
    print(f"Clients: {Client.query.count()} | Architects: {Architect.query.count()} | Companies: {Company.query.count()}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <path-to-xlsm>")
        sys.exit(1)

    with app.app_context():
        db.create_all()
        if User.query.count() == 0:
            print("No users exist yet — register a user via /register before signing in.")
        _warm_caches()
        import_workbook(sys.argv[1])


if __name__ == "__main__":
    main()
